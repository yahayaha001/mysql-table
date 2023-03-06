import hashlib
import math
import os
import re
import tkinter
import tkinter as tk
from tkinter import messagebox, filedialog
import threading
import pymysql
import tkinter as tk
from tkinter import ttk
import pandas as pd
from tkinter import filedialog
from openpyxl import Workbook
from tkinter import simpledialog
from tkinter import messagebox
from datetime import datetime
from login import run_login
from tqdm import tqdm

from sql_ui import login





# 连接 MySQL 数据库

def show_table_info(username):
# class MySQLTableInfo(tk.Frame):
    global current_user
    current_user = username
    def get_db_connection():
        conn = pymysql.connect(
            host='localhost',
            port=3306,
            user='gui',
            password='LBD8TrTBeMZFBa8t',
            db='gui',
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        return conn
    def close_connection(conn, cursor):
        cursor.close()
        conn.close()
        # print("MySQL connection is closed")


    # 创建 tkinter 窗口
    root = tk.Tk()
    root.title("MySQL Table Info")
    root.geometry("1580x600")


    style = ttk.Style()
    style.theme_use("default")

    style.configure("Custom.Treeview", font=("Calibri", 13), bordercolor='red',  borderwidth=1, relief="solid")

    tree = ttk.Treeview(root, style="Custom.Treeview", selectmode='browse', show='headings')
    tree.pack(side='left', fill='both', expand=True)





    table_header = ['ID','库名', '备注', '日期', '总量', '已搜索', '搜索有的', '待搜索', '异常', '账号归属']
    tree["columns"] = table_header
    # 设置第一列宽度为200
    # tree.column('#0', width=100)
    # 设置其它列宽度为100

    for col in table_header:

        if col=='日期':
            tree.column(col, width=180)
        elif col == 'id':
            tree.column(col, width=80)
        elif col == '库名':
            tree.column(col, width=180)
        elif col == '备注':
            tree.column(col, width=180)
        else:
            tree.column(col, width=70)

    # 表头
    for i in table_header:
        tree.column(i, anchor="center")
        tree.heading(i, text=i, anchor='center')

    # 查询所有表名
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor = conn.cursor()
    cursor.execute("show tables")
    tables = cursor.fetchall()
    close_connection(conn, cursor)
    # conn = get_db_connection()
    # cursor = conn.cursor()
    # 每页显示的条数
    page_size = 10

    # 获取数据总数
    total_count = len(tables)

    # 分页页数
    page_count = (total_count + page_size - 1) // page_size

    # 分页数据
    pages = [tables[i:i+page_size] for i in range(0, total_count, page_size)]

    # 当前页码
    current_page = 1

    # 显示指定页码的表格数据


    def is_super_user(username):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(f"SELECT is_suproot FROM users WHERE username = '{username}'")
        result = cursor.fetchone()
        close_connection(conn, cursor)
        if result and result['is_suproot'] == 1:
            return True
        else:
            return False

    def show_table_data(page_num, username):
        conn = get_db_connection()
        cursor = conn.cursor()
        tree.delete(*tree.get_children())
        # 查询ordea_all表中的数据
        if is_super_user(username):
            cursor.execute(
                f"SELECT * FROM ordea_all ORDER BY 日期 DESC LIMIT {page_size} OFFSET {(page_num - 1) * page_size}")
        else:
            cursor.execute(
                f"SELECT * FROM ordea_all WHERE 账号归属 = '{username}' ORDER BY 日期 DESC LIMIT {page_size} OFFSET {(page_num - 1) * page_size}")
        data = cursor.fetchall()
        for row in data:
            tree.insert('', 'end', values=(
                row['id'], row['库名'], row['备注'], row['日期'], row['总量'], row['已搜索'], row['搜索有'], row['待搜索'],row['异常'],row['账号归属']))
        close_connection(conn, cursor)

    # 显示第一页的数据
    show_table_data(current_page, username)

    # 分页标签
    page_label = tk.Label(root, text=f"页数： {current_page} / {page_count}")
    page_label.pack(side='bottom', padx=3, pady=3)


    def add_user():
        # 判断是否是超级管理员，非超级管理员不展示添加用户的功能
        if not is_super_user(username):
            return

        # 创建弹窗
        add_user_window = tk.Toplevel(root)
        add_user_window.title("添加用户")

        # 添加用户名、密码输入框和添加按钮
        username_label = tk.Label(add_user_window, text="用户名：")
        username_label.grid(row=0, column=0, padx=5, pady=5)
        username_entry = tk.Entry(add_user_window)
        username_entry.grid(row=0, column=1, padx=5, pady=5)

        password_label = tk.Label(add_user_window, text="密码：")
        password_label.grid(row=1, column=0, padx=5, pady=5)
        password_entry = tk.Entry(add_user_window, show="*")
        password_entry.grid(row=1, column=1, padx=5, pady=5)

        result_label = tk.Label(add_user_window, text="")
        result_label.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        def add_user_to_db():
            # 从输入框中获取用户名和密码
            new_username = username_entry.get()
            new_password = password_entry.get()

            # 对密码进行md5加密
            md5 = hashlib.md5()
            md5.update(new_password.encode())
            new_password_md5 = md5.hexdigest()

            # 查询数据库中是否已经有重名的用户
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(f"SELECT * FROM users WHERE username='{new_username}'")
            existing_user = cursor.fetchone()
            close_connection(conn, cursor)

            if existing_user:
                result_label.config(text=f"添加用户失败：用户名 '{new_username}' 已存在", fg="red")
            else:
                # 添加用户到users表中
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    cursor.execute(
                        f"INSERT INTO users (id, username, password) VALUES (NULL, '{new_username}', '{new_password_md5}')")
                    conn.commit()
                    close_connection(conn, cursor)
                    result_label.config(text="添加用户成功", fg="green")
                except Exception as e:
                    result_label.config(text=f"添加用户失败：{str(e)}", fg="red")
            refresh()
        add_button = tk.Button(add_user_window, text="添加", command=add_user_to_db)
        add_button.grid(row=3, column=1, padx=5, pady=5)

        # 将弹窗居中显示
        add_user_window.geometry("+%d+%d" % ((root.winfo_screenwidth() - add_user_window.winfo_reqwidth()) / 2,
                                             (root.winfo_screenheight() - add_user_window.winfo_reqheight()) / 2))

    def logout():
        conn = get_db_connection()
        cursor = conn.cursor()
        close_connection(conn, cursor)
        root.destroy()
        run_login()

    logout_button = tk.Button(root, text="退出登录", command=logout)
    logout_button.pack(side='left', padx=3, pady=3)


    if is_super_user(username):
        add_user_button = tk.Button(root, text="添加用户", command=add_user)
        add_user_button.pack(side='left', padx=3, pady=3)
# 上一页按钮
    def previous_page():
        global current_page
        if current_page > 1:
            current_page -= 1
            show_table_data(current_page,username)
            page_label.config(text=f"页数： {current_page} / {page_count}")

    previous_button = tk.Button(root, text="上一页", command=previous_page)
    previous_button.pack(side='bottom', padx=3, pady=3)

    # 下一页按钮
    def next_page():
        global current_page
        if current_page < page_count:
            current_page += 1
            show_table_data(current_page,username)
            page_label.config(text=f"页数： {current_page} / {page_count}")

    next_button = tk.Button(root, text="下一页", command=next_page)
    next_button.pack(side='bottom', padx=3, pady=3)




    # 创建添加表的对话框
    def create_table_dialog():
        dialog = tk.Toplevel(root)
        dialog.geometry("300x100")
        dialog.title("添加库")

        # 添加表名输入框
        label = tk.Label(dialog, text="输入库名")
        label.pack(side='top')
        entry = tk.Entry(dialog)
        entry.pack(side='top')

        # 添加确认按钮
        def confirm():
            table_name = entry.get()
            if not table_name:
                return
            create_table(username,table_name)
            dialog.destroy()

        refresh()
        confirm_button = tk.Button(dialog, text="确认", command=confirm)
        confirm_button.pack(side='bottom',padx=3, pady=3)







    # 删除库
    def delete_table():
        # 获取选中的表名

        conn = get_db_connection()
        cursor = conn.cursor()
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("错误", "未选中任何表格！")
            return
        table_name = tree.item(selected_item)['values'][1]
        if not table_name.startswith('ordea_'):
            messagebox.showerror("错误", "无法删除此表格！")
            return
        # 弹出确认删除对话框
        if not tkinter.messagebox.askyesno("确认", f"是否确定删除 '{table_name}'？"):
            return
        # 删除 ordea_all 表中对应的记录
        try:
            cursor.execute(f"DELETE FROM ordea_all WHERE 库名='{table_name}'")
            conn.commit()
        except Exception as e:
            print(e)
            conn.rollback()
            messagebox.showerror("错误", "删除记录失败！")
            return
        # 弹出是否删除数据表的对话框
        if tkinter.messagebox.askyesno("确认", f"订单删除成功，是否也请客 '{table_name}'所有数据？"):
            # 删除数据表
            try:
                cursor.execute(f"DROP TABLE `{table_name}`")
                conn.commit()
            except Exception as e:
                print(e)
                conn.rollback()
                messagebox.showerror("错误", "删除数据表失败！")
                return
        # 刷新表格内容
        close_connection(conn, cursor)
        refresh()



    # 创建删除库按钮
    delete_button = tk.Button(root, text="删除库",  command=delete_table)
    delete_button.pack(side='bottom', padx=3, pady=3)



# 创建导出按钮
    def export_to_excel_yss():
        conn = get_db_connection()
        cursor = conn.cursor()
        selected_items = tree.selection()
        if not selected_items:
            return

        file_name = ""
        for item in selected_items:
            table_name = tree.set(item, '库名')

            cursor.execute(f"select count(*) from {table_name}")
            result = cursor.fetchone()
            if result:
                count = result['count(*)']
                table_name1 = table_name.replace("ordea_", "")
                file_name += f"{table_name1}_所有_{count}.xlsx"
                break

        file_path = filedialog.asksaveasfilename(initialfile=file_name, defaultextension='.xlsx')
        if not file_path:
            return

        progress_window = tk.Toplevel(root)
        progress_window.title("导出进度")
        progress_label = tk.Label(progress_window, text="正在导出中，请稍等...")
        progress_label.pack()
        progress_bar = tk.ttk.Progressbar(progress_window, mode="indeterminate")
        progress_bar.pack(pady=10)
        progress_bar.start(10)
        progress_window.update()

        wb = Workbook()
        ws = wb.active
        headers = ['id', '手机号', '名字', '性别', '昵称', '账号', '省份', '城市', '会员', '头像', '归属省份', '归属城市',
                   '区划代码', '运营商', '状态']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)

        for item in selected_items:
            table_name = tree.set(item, '库名')
            cursor.execute(f"select * from {table_name} ")
            data = cursor.fetchall()
            for i, row in enumerate(data, start=2):
                for j, value in enumerate(row.values()):
                    ws.cell(row=i, column=j + 1, value=value)
        if os.path.exists(file_path):
            try:
                with open(file_path, 'wb') as f:
                    f.write('test'.encode('utf-8'))
            except PermissionError:
                progress_window.destroy()
                messagebox.showerror("错误", "文件被占用，无法导出，请关闭该文件后重试")
                return
        wb.save(file_path)
        close_connection(conn, cursor)
        progress_bar.stop()
        progress_window.destroy()

        messagebox.showinfo("导出完成", f"导出成功，共导出{count}条数据")
        refresh()


    export_button = tk.Button(root, text="导出所有", command=export_to_excel_yss)
    export_button.pack(side='bottom', padx=3, pady=3)


    def export_to_excel_ssyd():
        conn = get_db_connection()
        cursor = conn.cursor()
        selected_items = tree.selection()
        if not selected_items:
            return

        file_name = ""
        for item in selected_items:
            table_name = tree.set(item, '库名')

            cursor.execute(f"select count(*) from {table_name} WHERE `gender` IS NOT NULL AND `gender` != ''")
            result = cursor.fetchone()
            if result:
                count = result['count(*)']
                table_name1 = table_name.replace("ordea_", "")
                file_name += f"{table_name1}_搜索有的_{count}.xlsx"
                break

        file_path = filedialog.asksaveasfilename(initialfile=file_name, defaultextension='.xlsx')
        if not file_path:
            return

        progress_window = tk.Toplevel(root)
        progress_window.title("导出进度")
        progress_label = tk.Label(progress_window, text="正在导出中，请稍等...")
        progress_label.pack()
        progress_bar = tk.ttk.Progressbar(progress_window, mode="indeterminate")
        progress_bar.pack(pady=10)
        progress_bar.start(10)
        progress_window.update()

        wb = Workbook()
        ws = wb.active
        headers = ['id', '手机号', '名字', '性别', '昵称', '账号', '省份', '城市', '会员', '头像', '归属省份', '归属城市',
                   '区划代码', '运营商', '状态']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)

        for item in selected_items:
            table_name = tree.set(item, '库名')
            cursor.execute(f"select * from {table_name} WHERE `gender` IS NOT NULL AND `gender` != '' ")
            data = cursor.fetchall()
            for i, row in enumerate(data, start=2):
                for j, value in enumerate(row.values()):
                    ws.cell(row=i, column=j + 1, value=value)
        if os.path.exists(file_path):
            try:
                with open(file_path, 'wb') as f:
                    f.write('test'.encode('utf-8'))
            except PermissionError:
                progress_window.destroy()
                messagebox.showerror("错误", "文件被占用，无法导出，请关闭该文件后重试")
                return
        wb.save(file_path)
        close_connection(conn, cursor)
        progress_bar.stop()
        progress_window.destroy()

        messagebox.showinfo("导出完成", f"导出成功，共导出{count}条数据")
        refresh()


    export_button = tk.Button(root, text="导出搜索有的", command=export_to_excel_ssyd)
    export_button.pack(side='bottom', padx=3, pady=3)


    def export_to_excel_dss():
        conn = get_db_connection()
        cursor = conn.cursor()
        selected_items = tree.selection()
        if not selected_items:
            return

        file_name = ""
        for item in selected_items:
            table_name = tree.set(item, '库名')

            cursor.execute(f"select count(*) from {table_name}  where status = 0 or status = 2 or status = 3")
            result = cursor.fetchone()
            if result:
                count = result['count(*)']
                table_name1 = table_name.replace("ordea_", "")
                file_name += f"{table_name1}_待搜索_{count}.xlsx"
                break

        file_path = filedialog.asksaveasfilename(initialfile=file_name, defaultextension='.xlsx')
        if not file_path:
            return

        progress_window = tk.Toplevel(root)
        progress_window.title("导出进度")
        progress_label = tk.Label(progress_window, text="正在导出中，请稍等...")
        progress_label.pack()
        progress_bar = tk.ttk.Progressbar(progress_window, mode="indeterminate")
        progress_bar.pack(pady=10)
        progress_bar.start(10)
        progress_window.update()

        wb = Workbook()
        ws = wb.active
        headers = ['id', '手机号', '名字', '性别', '昵称', '账号', '省份', '城市', '会员', '头像', '归属省份', '归属城市',
                   '区划代码', '运营商', '状态']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)

        for item in selected_items:
            table_name = tree.set(item, '库名')
            cursor.execute(f"select * from {table_name}  where status = 0 or status = 2 or status = 3")
            data = cursor.fetchall()
            for i, row in enumerate(data, start=2):
                for j, value in enumerate(row.values()):
                    ws.cell(row=i, column=j + 1, value=value)
        if os.path.exists(file_path):
            try:
                with open(file_path, 'wb') as f:
                    f.write('test'.encode('utf-8'))
            except PermissionError:
                progress_window.destroy()
                messagebox.showerror("错误", "文件被占用，无法导出，请关闭该文件后重试")
                return
        wb.save(file_path)
        close_connection(conn, cursor)
        progress_bar.stop()
        progress_window.destroy()

        messagebox.showinfo("导出完成", f"导出成功，共导出{count}条数据")
        refresh()


    export_button = tk.Button(root, text="导出待搜索", command=export_to_excel_dss)
    export_button.pack(side='bottom', padx=3, pady=3)


    def export_to_excel_yc():
        conn = get_db_connection()
        cursor = conn.cursor()
        selected_items = tree.selection()
        if not selected_items:
            return

        file_name = ""
        for item in selected_items:
            table_name = tree.set(item, '库名')

            cursor.execute(f"select count(*) from {table_name}  where status = 3 or status = 2")
            result = cursor.fetchone()
            if result:
                count = result['count(*)']
                table_name1 = table_name.replace("ordea_", "")
                file_name += f"{table_name1}_异常_{count}.xlsx"
                break

        file_path = filedialog.asksaveasfilename(initialfile=file_name, defaultextension='.xlsx')
        if not file_path:
            return

        progress_window = tk.Toplevel(root)
        progress_window.title("导出进度")
        progress_label = tk.Label(progress_window, text="正在导出中，请稍等...")
        progress_label.pack()
        progress_bar = tk.ttk.Progressbar(progress_window, mode="indeterminate")
        progress_bar.pack(pady=10)
        progress_bar.start(10)
        progress_window.update()

        wb = Workbook()
        ws = wb.active
        headers = ['id', '手机号', '名字', '性别', '昵称', '账号', '省份', '城市', '会员', '头像', '归属省份', '归属城市',
                   '区划代码', '运营商', '状态']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)

        for item in selected_items:
            table_name = tree.set(item, '库名')
            cursor.execute(f"select * from {table_name}  where status = 3 or status = 2")
            data = cursor.fetchall()
            for i, row in enumerate(data, start=2):
                for j, value in enumerate(row.values()):
                    ws.cell(row=i, column=j + 1, value=value)
        if os.path.exists(file_path):
            try:
                with open(file_path, 'wb') as f:
                    f.write('test'.encode('utf-8'))
            except PermissionError:
                progress_window.destroy()
                messagebox.showerror("错误", "文件被占用，无法导出，请关闭该文件后重试")
                return
        wb.save(file_path)
        close_connection(conn, cursor)
        progress_bar.stop()
        progress_window.destroy()

        messagebox.showinfo("导出完成", f"导出成功，共导出{count}条数据")
        refresh()


    export_button = tk.Button(root, text="导出异常", command=export_to_excel_yc)
    export_button.pack(side='bottom', padx=3, pady=3)
    def export_to_excel_yss():
        conn = get_db_connection()
        cursor = conn.cursor()
        selected_items = tree.selection()
        if not selected_items:
            return

        file_name = ""
        for item in selected_items:
            table_name = tree.set(item, '库名')

            cursor.execute(f"select count(*) from {table_name} where status = 1")
            result = cursor.fetchone()
            if result:
                count = result['count(*)']
                table_name1 = table_name.replace("ordea_", "")
                file_name += f"{table_name1}_已搜索_{count}.xlsx"
                break

        file_path = filedialog.asksaveasfilename(initialfile=file_name, defaultextension='.xlsx')
        if not file_path:
            return

        progress_window = tk.Toplevel(root)
        progress_window.title("导出进度")
        progress_label = tk.Label(progress_window, text="正在导出中，请稍等...")
        progress_label.pack()
        progress_bar = tk.ttk.Progressbar(progress_window, mode="indeterminate")
        progress_bar.pack(pady=10)
        progress_bar.start(10)
        progress_window.update()

        wb = Workbook()
        ws = wb.active
        headers = ['id', '手机号', '名字', '性别', '昵称', '账号', '省份', '城市', '会员', '头像', '归属省份', '归属城市',
                   '区划代码', '运营商', '状态']
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)

        for item in selected_items:
            table_name = tree.set(item, '库名')
            cursor.execute(f"select * from {table_name} where status = 1")
            data = cursor.fetchall()
            for i, row in enumerate(data, start=2):
                for j, value in enumerate(row.values()):
                    ws.cell(row=i, column=j + 1, value=value)
        if os.path.exists(file_path):
            try:
                with open(file_path, 'wb') as f:
                    f.write('test'.encode('utf-8'))
            except PermissionError:
                progress_window.destroy()
                messagebox.showerror("错误", "文件被占用，无法导出，请关闭该文件后重试")
                return
        wb.save(file_path)
        close_connection(conn, cursor)
        progress_bar.stop()
        progress_window.destroy()

        messagebox.showinfo("导出完成", f"导出成功，共导出{count}条数据")
        refresh()

    export_button = tk.Button(root, text="导出已搜索", command=export_to_excel_yss)
    export_button.pack(side='bottom', padx=3, pady=3)


    def reset_data():
        conn = get_db_connection()
        cursor = conn.cursor()
        selected_items = tree.selection()
        if not selected_items:
            return

        reset_count = 0
        for item in selected_items:
            table_name = tree.set(item, '库名')
            cursor.execute(f"update {table_name} set status=0 where status=2 or status=3")
            reset_count += cursor.rowcount
            conn.commit()

        close_connection(conn, cursor)
        messagebox.showinfo("提示", f"本次成功重置({reset_count})条异常数据")
        refresh()


    reset_button = tk.Button(root, text="重置异常", command=reset_data)
    reset_button.pack(side='bottom', padx=3, pady=3)


    def reset_data_kb():
        conn = get_db_connection()
        cursor = conn.cursor()
        selected_items = tree.selection()
        if not selected_items:
            return

        reset_count = 0
        for item in selected_items:
            table_name = tree.set(item, '库名')
            cursor.execute(f"update {table_name} set status=0 WHERE `gender` IS NOT NULL AND `gender` != '' ")
            reset_count += cursor.rowcount
            conn.commit()

        close_connection(conn, cursor)
        messagebox.showinfo("提示", f"本次成功重置({reset_count})条空白数据")
        refresh()


    reset_button = tk.Button(root, text="重置空白", command=reset_data_kb)
    reset_button.pack(side='bottom', padx=3, pady=3)

# 新增“修改备注”的对话框
    def modify_remark_dialog():
        conn = get_db_connection()
        cursor = conn.cursor()
        # 获取当前选中行的数据
        selected_item = tree.focus()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a row first.")
            return
        values = tree.item(selected_item, 'values')
        table_name = values[1]
        remark = values[1]

        dialog = tk.Toplevel(root)
        dialog.geometry("300x200")
        dialog.title("修改备注")

        # 显示库名和当前备注信息
        label1 = tk.Label(dialog, text=f"库名：{table_name}")
        label1.pack(side='top')
        label2 = tk.Label(dialog, text="当前备注：")
        label2.pack(side='top')
        old_remark = tk.Label(dialog, text=remark)
        old_remark.pack(side='top')

        # 添加输入框和确认按钮
        label3 = tk.Label(dialog, text="新备注：")
        label3.pack(side='top')
        entry = tk.Entry(dialog)
        entry.pack(side='top')

        def confirm():
            conn = get_db_connection()
            cursor = conn.cursor()
            new_remark = entry.get()
            if not new_remark:
                messagebox.showwarning("Warning", "Please enter new remark.")
                return
            # 更新数据库中的备注信息
            cursor.execute(f"UPDATE ordea_all SET 备注='{new_remark}' WHERE 库名='{table_name}'")
            conn.commit()
            # 更新表格中的备注信息
            tree.item(selected_item, values=(table_name, new_remark, values[2]))
            dialog.destroy()
            refresh()
        close_connection(conn, cursor)

        confirm_button = tk.Button(dialog, text="确认", command=confirm)
        confirm_button.pack(side='bottom',padx=3, pady=3)


    # 新增“修改备注”按钮
    modify_remark_button = tk.Button(root, text="修改备注", command=modify_remark_dialog)
    modify_remark_button.pack(side='bottom',padx=3, pady=3)


    def import_data():
        conn = get_db_connection()
        cursor = conn.cursor()
        # 获取选中的表名
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("错误", "未选中任何表格！")
            return
        table_name = tree.item(selected_item)['values'][1]
        if not table_name.startswith('ordea_'):
            messagebox.showerror("错误", "无法导入数据到此表格！")
            return

        # 弹出文件选择对话框
        file_path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        if not file_path:
            return

        # 弹出导入进度对话框
        import_progress = tk.Toplevel(root)
        import_progress.title("导入进度")
        import_progress.geometry("400x100")  # 设置弹窗大小
        progress_label = tk.Label(import_progress, text="正在导入数据，请稍等...", font=("Helvetica", 16))
        progress_label.pack(pady=20)

        def update_progress(current_count, total_count, progress_label):
            progress_percent = current_count * 100 / total_count
            progress_label.config(text=f"已导入 {current_count} / {total_count} 行数据 ({progress_percent:.2f}%)")
            import_progress.update_idletasks()


        def import_thread():
            # 打开文件并读取数据
            with open(file_path, 'r') as f:
                lines = f.readlines()
                total_count = len(lines)
                current_count = 0
                success_count = 0
                repeat_count = 0
                # 分批导入数据
                batch_size = 10000
                for i in range(0, total_count, batch_size):
                    batch_data = []
                    for line in lines[i:i + batch_size]:
                        data = line.strip().split()
                        batch_data.append(data)

                    # 插入数据到数据库
                    try:
                        cursor.executemany(f"INSERT IGNORE INTO `{table_name}` (`phone`, `status`) VALUES (%s, 0)",
                                           batch_data)
                        conn.commit()
                        success_count += cursor.rowcount
                        repeat_count += len(batch_data) - cursor.rowcount
                    except Exception as e:
                        print(e)
                        conn.rollback()
                        messagebox.showerror("错误", "插入数据失败，请检查是否非11位手机号。")
                        close_connection(conn, cursor)
                        return
                    refresh
                    # 更新导入进度
                    current_count += len(batch_data)
                    progress_percent = current_count * 100 / total_count
                    progress_label.config(
                        text=f"已导入 {current_count} / {total_count} 行数据 ({progress_percent:.2f}%)")
                    import_progress.update_idletasks()

                # 关闭导入进度对话框
                import_progress.destroy()
                messagebox.showinfo("信息", f"数据导入完成！成功导入 {success_count} 行数据，重复数据 {repeat_count} 行。")
                close_connection(conn, cursor)
            refresh()
        # import_thread = threading.Thread(target=import_thread)
        # import_thread.start()
        import_thread()
    # 创建导入数据的按钮
    import_button = tk.Button(root, text="导入数据", command=import_data)
    import_button.pack(side='bottom', padx=3, pady=3)



    def open_upload_dialog():
        conn = get_db_connection()
        cursor = conn.cursor()
        # 获取选中的表名
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showerror("错误", "未选中任何表格！")
            return
        table_name = tree.item(selected_item)['values'][1]
        if not table_name.startswith('ordea_'):
            messagebox.showerror("错误", "无法导入数据到此表格！")
            return

        # 弹出文本输入框
        upload_window = tk.Toplevel(root)
        upload_window.title("输入上传数据")
        upload_text = tk.Text(upload_window, width=50, height=10)
        upload_text.pack(padx=10, pady=10)

        def import_data_from_text():
            # 读取输入的数据
            data = upload_text.get("1.0", "end").strip().split()
            if not data:
                messagebox.showwarning("警告", "输入数据为空！")
                return

            # 弹出导入进度对话框
            import_progress = tk.Toplevel(root)
            import_progress.title("导入进度")
            progress_label = tk.Label(import_progress, text="正在导入数据，请稍等...")
            progress_label.pack()

            # 插入数据到数据库
            try:
                cursor.executemany(f"INSERT IGNORE INTO `{table_name}` (`phone`, `status`) VALUES (%s, 0)", data)
                conn.commit()
            except Exception as e:
                print(e)
                conn.rollback()
                messagebox.showerror("错误", "插入数据失败,请检查是否数据重复，或者非11位手机号。")
                return

            # 关闭导入进度对话框
            import_progress.destroy()
            messagebox.showinfo("信息", "数据导入完成！")

            # 关闭输入上传数据的对话框
            upload_window.destroy()

            close_connection(conn, cursor)
            refresh()
        # 添加导入按钮
        import_button = tk.Button(upload_window, text="导入", command=import_data_from_text)
        import_button.pack(pady=10)

    # 创建上传数据按钮
    upload_button = tk.Button(root, text="输入数据", command=open_upload_dialog)
    upload_button.pack(side='bottom',padx=3, pady=3)


    def create_table(username, table_name):
        conn = get_db_connection()
        cursor = conn.cursor()
        # 查询表名是否已经存在
        cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        result = cursor.fetchone()
        if result:
            messagebox.showerror("Error", "Table already exists!")
            return
        table_name = 'ordea_' + table_name
        # 创建表
        sql = f'''
                    CREATE TABLE `{table_name}` (
                      `id` INT NOT NULL AUTO_INCREMENT,
                      `phone` VARCHAR(11) NOT NULL,
                      `name` VARCHAR(255),
                      `gender` VARCHAR(255),
                      `nickname` VARCHAR(255),
                      `account` VARCHAR(255),
                      `province` VARCHAR(255),
                      `city` VARCHAR(255),
                      `member` VARCHAR(255),
                      `avatar` VARCHAR(255),
                      `province_isp` VARCHAR(255),
                      `city_isp` VARCHAR(255),
                      `zoning_code` VARCHAR(255),
                      `isp` VARCHAR(255),
                      `status` INT,
                      `by1` VARCHAR(255),
                      `by2` VARCHAR(255),
                      `by3` VARCHAR(255),
                      PRIMARY KEY (`id`),
                      UNIQUE KEY `unique_phone` (`phone`)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
            '''
        try:
            cursor.execute(sql.format(table_name=table_name))
            print('1')
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute(
                f"INSERT INTO ordea_all (库名, 备注, 日期, 账号归属) VALUES ('{table_name}', '', '{now}', '{username}')")
            conn.commit()
            tree.insert('', 'end', values=(table_name, 0))
            show_table_data(current_page,username)  # 刷新整个表格的内容显示
        except Exception as e:
            print(e)
            conn.rollback()
            messagebox.showerror("Error", "Create table failed!")

        close_connection(conn, cursor)
        refresh()

    create_table_button = tk.Button(root, text="添加库", command=create_table_dialog)
    create_table_button.pack(side='bottom', padx=3, pady=3)


    # 刷新表格内容
    def refresh_table_data():
        global tables
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM ordea_all")
        tables = cursor.fetchall()
        tree.delete(*tree.get_children())
        print(tables)
        for table in tables:
            id = table['id']
            table_name = table['库名']
            remark = table['备注']
            date = table['日期']
            cursor.execute(f"select count(*) from `{table_name}`")
            count = cursor.fetchone()['count(*)']
            tree.insert('', 'end', values=(id,table_name, remark, date, count))
        # 更新分页相关数据
        global total_count, page_count, pages
        total_count = len(tables)
        page_count = (total_count + page_size - 1) // page_size
        pages = [tables[i:i+page_size] for i in range(0, total_count, page_size)]
        # 显示第一页的数据
        global current_page
        current_page = 1
        show_table_data(current_page,username)
        # 更新分页标签
        page_label.config(text=f"分页： {current_page} / {page_count}")
        close_connection(conn, cursor)




    # 新增“刷新”按钮
    def refresh():
        # 清空表格
        conn = get_db_connection()
        cursor = conn.cursor()
        tree.delete(*tree.get_children())
        # 查询数据库
        cursor.execute("SELECT * FROM ordea_all ORDER BY 日期 DESC")
        data = cursor.fetchall()
        if not data:
            return
        for row in data:
            table_name = row['库名']
            cursor.execute(f"SELECT COUNT(*) FROM `{table_name}`")
            result = cursor.fetchone()
            # print(f"Result: {result}")

            count = result.get('COUNT(*)', 0)

            cursor.execute(f"SELECT COUNT(*) FROM `{table_name}` WHERE `status` IN (0, 2)")
            result = cursor.fetchone()
            pending_count = result.get('COUNT(*)', 0)

            cursor.execute(f"SELECT COUNT(*) FROM `{table_name}` WHERE `status` = 1")
            result = cursor.fetchone()
            searched_count = result.get('COUNT(*)', 0)

            cursor.execute(f"SELECT COUNT(*) FROM `{table_name}` WHERE `status` = 3")
            result = cursor.fetchone()
            exception_count = result.get('COUNT(*)', 0)

            cursor.execute(f"SELECT COUNT(*) FROM `{table_name}` WHERE `gender` IS NOT NULL AND `gender` != ''")
            result = cursor.fetchone()
            search_success_count = result.get('COUNT(*)', 0)

            cursor.execute(f"UPDATE ordea_all SET `总量`={count}, `待搜索`={pending_count}, `已搜索`={searched_count}, `异常`={exception_count}, `搜索有`={search_success_count} WHERE `库名`='{table_name}'")
            conn.commit()

            # tree.insert('', 'end', values=(row['库名'], row['备注'], count, pending_count, searched_count, exception_count, search_success_count, row['日期']))

        # 更新页码和分页数据
        global total_count, page_count, pages
        total_count = len(data)
        page_count = (total_count + page_size - 1) // page_size
        pages = [data[i:i+page_size] for i in range(0, total_count, page_size)]
        # 更新当前页码和分页标签
        global current_page
        current_page = 1
        page_label.config(text=f"页数： {current_page} / {page_count}")

        # 显示第一页的数据
        show_table_data(current_page,username)
        close_connection(conn, cursor)


    refresh_button = tk.Button(root, text="刷新", command=refresh)
    refresh_button.pack(side='bottom',padx=3, pady=3)




def run_login(master=None):
    # 创建登录窗口
    login_window = tk.Tk()
    login_window.geometry("300x200")
    login_window.title("登录")

    # 创建用户名和密码输入框
    username_label = tk.Label(login_window, text="用户名：")
    username_label.pack(pady=10)
    username_entry = tk.Entry(login_window)
    username_entry.pack(pady=5)

    password_label = tk.Label(login_window, text="密码：")
    password_label.pack()
    password_entry = tk.Entry(login_window, show="*")
    password_entry.pack(pady=5)

    # 定义登录按钮的点击事件
    def login():
        # 获取输入的用户名和密码
        username = username_entry.get()
        password = password_entry.get()
        hashed_password = hashlib.md5(password.encode('utf-8')).hexdigest()
        # 连接数据库
        conn = pymysql.connect(
            host='localhost',
            port=3306,
            user='gui',
            password='LBD8TrTBeMZFBa8t',
            db='gui',
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        cursor = conn.cursor()

        # 查询用户名和密码是否正确
        cursor.execute(f"SELECT * FROM users WHERE username='{username}' AND password='{hashed_password}'")
        result = cursor.fetchone()
        if result:
            # 用户名和密码验证通过，关闭登录窗口并执行回调函数
            login_window.destroy()
            login_success(result['username'])
        else:
            # 用户名和密码验证失败，弹出提示框
            messagebox.showerror("错误", "用户名或密码不正确！")

        cursor.close()
        conn.close()

    # 创建登录按钮
    login_button = tk.Button(login_window, text="登录", command=login)
    login_button.pack(pady=10)

    login_window.mainloop()


def login_success(username):
    # messagebox.showinfo("提示", f"欢迎 {username} 登录！")
    show_table_info(username)


if __name__ == '__main__':
    # 直接打开登录窗口
    run_login()
    # 启动 Tkinter 事件循环
